"""
用户模块视图
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.forms import UserCreationForm, PasswordChangeForm
from django.http import JsonResponse
from django.utils import timezone
from django.db.models import Q
from datetime import timedelta
from .models import User, Department, Role
from .forms import CustomAuthenticationForm, UserForm, DepartmentForm, RoleForm, UserProfileForm


def is_admin(user):
    """检查用户是否为管理员或试验室主管"""
    return user.role in ['admin', 'manager']


def admin_required(view_func):
    """
    自定义装饰器：检查用户是否为管理员或试验室主管
    如果不是，显示权限不足的提示而不是跳转到登录页面
    """
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, '请先登录')
            return redirect('users:login')
        if not is_admin(request.user):
            messages.error(request, '权限不足：只有试验室主管和系统管理员才能执行此操作')
            # 重定向到用户列表页面而不是登录页面
            return redirect('users:user_list')
        return view_func(request, *args, **kwargs)
    return _wrapped_view


def user_login(request):
    """用户登录"""
    # 即使用户已登录，也显示登录页面
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            # 根据用户角色重定向到不同页面
            if user.role in ['guest', 'engineer']:
                # 访客和试验工程师跳转到试验任务仪表盘
                return redirect('tasks:dashboard')
            else:
                return redirect('dashboard')
        else:
            # 显示表单中的具体错误信息，而不是统一的"用户名或密码错误"
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, error)
    else:
        form = CustomAuthenticationForm()
    
    return render(request, 'users/login.html', {'form': form})


def user_logout(request):
    """用户登出"""
    logout(request)
    messages.info(request, '您已成功退出登录')
    return redirect('users:login')


@login_required
def user_profile(request):
    """用户个人资料"""
    return render(request, 'users/profile.html', {'user': request.user})


@login_required
def edit_profile(request):
    """编辑个人资料"""
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            user = form.save()
            # 如果修改了密码，需要重新登录以保持会话
            if form.cleaned_data.get('new_password'):
                update_session_auth_hash(request, user)
                messages.success(request, '个人资料及密码更新成功！')
            else:
                messages.success(request, '个人资料更新成功！')
            return redirect('users:profile')
        else:
            messages.error(request, '更新失败，请检查输入信息')
    else:
        form = UserProfileForm(instance=request.user)
    
    return render(request, 'users/edit_profile.html', {'form': form, 'user': request.user})


@login_required
def change_password(request):
    """修改密码"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # 防止密码修改后session失效
            messages.success(request, '密码修改成功！')
            return redirect('users:profile')
        else:
            messages.error(request, '密码修改失败，请检查输入信息')
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'users/password_change.html', {
        'form': form,
        'title': '修改密码'
    })


@login_required
def user_list(request):
    """用户列表"""
    # 统计信息
    # 为避免静态分析工具报错，使用更明确的方式
    total_users = User.objects.count()
    active_users = User.objects.filter(account_status=True).count()
    staff_users = User.objects.filter(is_staff=True).count()
    departments_count = Department.objects.count()
    
    stats = {
        'total_users': total_users,
        'active_users': active_users,
        'staff_users': staff_users,
        'departments_count': departments_count,
    }

    # 搜索和筛选
    search_query = request.GET.get('search', '')
    department_filter = request.GET.get('department', '')
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')

    users = User.objects.all().order_by('-date_joined')

    if search_query:
        # 为避免静态分析工具报错，使用列表和循环方式构建查询
        # 移除了first_name和last_name的搜索字段
        search_fields = [
            'username__icontains',
            'employee_id__icontains',
            'email__icontains'
        ]
        
        q_objects = Q()
        for field in search_fields:
            q_objects.add(Q(**{field: search_query}), Q.OR)
        users = users.filter(q_objects)

    if department_filter:
        users = users.filter(department_id=department_filter)
        
    if role_filter:
        users = users.filter(role=role_filter)
        
    if status_filter:
        if status_filter == 'active':
            users = users.filter(account_status=True)
        elif status_filter == 'inactive':
            users = users.filter(account_status=False)

    # 分页
    from django.core.paginator import Paginator
    paginator = Paginator(users, 10)  # 每页10个用户
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 为避免静态分析工具报错，使用更明确的方式获取部门列表
    departments = Department.objects.all()
    roles = User.USER_ROLE_CHOICES

    return render(request, 'users/user_list.html', {
        'page_obj': page_obj,
        'stats': stats,
        'departments': departments,
        'roles': roles,
        'search_query': search_query,
        'department_filter': department_filter,
        'role_filter': role_filter,
        'status_filter': status_filter,
    })


@login_required
def user_detail(request, user_id):
    """用户详情"""
    user_detail = get_object_or_404(User, id=user_id)
    return render(request, 'users/user_detail.html', {'user_detail': user_detail})


@login_required
@admin_required
def user_create(request):
    """创建用户（管理员）"""
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            # 设置密码
            password = form.cleaned_data.get('password')
            if password:
                user.set_password(password)
            else:
                # 如果没有提供密码，设置默认密码
                user.set_password('123456')
            user.save()
            messages.success(request, '用户创建成功！')
            return redirect('users:user_list')
        else:
            messages.error(request, '创建失败，请检查输入信息')
    else:
        form = UserForm()

    return render(request, 'users/user_form.html', {
        'form': form,
        'title': '添加用户'
    })


@login_required
@admin_required
def user_edit(request, user_id):
    """编辑用户（管理员）"""
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            user = form.save(commit=False)
            # 如果输入了新密码，则更新密码
            password = form.cleaned_data.get('password')
            if password:
                user.set_password(password)
            user.save()
            messages.success(request, '用户信息更新成功！')
            return redirect('users:user_list')
        else:
            messages.error(request, '更新失败，请检查输入信息')
    else:
        form = UserForm(instance=user)

    return render(request, 'users/user_form.html', {
        'form': form,
        'title': '编辑用户',
        'user': user
    })


@login_required
@admin_required
def user_delete(request, user_id):
    """删除用户（管理员）"""
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        # 不允许删除管理员自身
        if user.is_staff and user != request.user:
            messages.error(request, '不允许删除管理员！')
        else:
            user.delete()
            messages.success(request, '用户删除成功！')

        return redirect('users:user_list')

    return render(request, 'users/user_confirm_delete.html', {'user': user})


@login_required
@admin_required
def toggle_user_status(request, user_id):
    """切换用户状态（激活/禁用）"""
    if request.method == 'POST' and request.is_ajax():
        user = get_object_or_404(User, id=user_id)
        # 切换账户状态
        user.account_status = not user.account_status
        user.save()

        return JsonResponse({
            'success': True,
            'is_active': user.account_status,
            'message': '用户状态更新成功！'
        })

    return JsonResponse({
        'success': False,
        'message': '请求方法不正确'
    })


@login_required
@admin_required
def export_users(request):
    """导出用户数据到Excel"""
    # 获取搜索和筛选参数
    search_query = request.GET.get('search', '')
    department_filter = request.GET.get('department', '')
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')

    # 构建查询
    users = User.objects.all().order_by('-date_joined')

    if search_query:
        search_fields = [
            'username__icontains',
            'employee_id__icontains',
            'email__icontains'
        ]
        
        q_objects = Q()
        for field in search_fields:
            q_objects.add(Q(**{field: search_query}), Q.OR)
        users = users.filter(q_objects)

    if department_filter:
        users = users.filter(department_id=department_filter)
        
    if role_filter:
        users = users.filter(role=role_filter)
        
    if status_filter:
        if status_filter == 'active':
            users = users.filter(account_status=True)
        elif status_filter == 'inactive':
            users = users.filter(account_status=False)

    # 创建Excel文件
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    from django.utils import timezone
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "用户数据"
    
    # 设置标题行样式
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加标题行
    headers = ['工号', '用户名', '邮箱', '部门', '角色', '手机号', '职位', '状态', '创建时间']
    ws.append(headers)
    
    # 设置标题行样式
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 角色映射
    role_mapping = dict(User.USER_ROLE_CHOICES)
    
    # 添加数据行
    for user in users:
        department_name = user.department.name if user.department else ''
        role_name = role_mapping.get(user.role, user.role)
        status_name = '在职' if user.account_status else '离职'
        
        # 正确处理时区并格式化时间
        if user.date_joined:
            # 将时间转换为本地时区
            local_time = timezone.localtime(user.date_joined)
            created_time = local_time.strftime('%Y-%m-%d %H:%M:%S')
        else:
            created_time = ''
        
        row = [
            user.employee_id,
            user.username,
            user.email or '',
            department_name,
            role_name,
            user.phone or '',
            user.position or '',
            status_name,
            created_time
        ]
        ws.append(row)
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 将Excel文件保存到内存中
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # 创建HttpResponse对象
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="users_export.xlsx"'
    
    return response


@login_required
@admin_required
def department_list(request):
    """部门列表"""
    # 搜索
    search_query = request.GET.get('search', '')
    
    departments = Department.objects.all().order_by('name')
    
    if search_query:
        departments = departments.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    # 分页
    from django.core.paginator import Paginator
    paginator = Paginator(departments, 10)  # 每页10个部门
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'users/department_list.html', {
        'departments': page_obj,
        'search_query': search_query,
        'is_paginated': page_obj.has_other_pages(),
        'page_obj': page_obj,
    })


@login_required
@admin_required
def department_create(request):
    """创建部门"""
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '部门创建成功！')
            return redirect('users:department_list')
    else:
        form = DepartmentForm()
    
    return render(request, 'users/department_form.html', {
        'form': form,
        'title': '添加部门'
    })


@login_required
@admin_required
def department_update(request, pk):
    """更新部门"""
    department = get_object_or_404(Department, pk=pk)
    
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            messages.success(request, '部门信息更新成功！')
            return redirect('users:department_list')
    else:
        form = DepartmentForm(instance=department)
    
    return render(request, 'users/department_form.html', {
        'form': form,
        'title': '编辑部门',
        'department': department
    })


@login_required
@admin_required
def department_detail(request, pk):
    """部门详情"""
    department = get_object_or_404(Department, pk=pk)
    # 获取该部门的用户
    users = User.objects.filter(department=department).order_by('username')
    
    return render(request, 'users/department_detail.html', {
        'department': department,
        'users': users
    })


# ==================== 角色管理相关视图 ====================

@login_required
@admin_required
def role_list(request):
    """角色列表（仅系统管理员可访问）"""
    # 只有系统管理员可以管理角色
    if request.user.role != 'admin':
        messages.error(request, '权限不足：只有系统管理员才能管理角色！')
        return redirect('index')
    
    # 搜索
    search_query = request.GET.get('search', '')
    
    roles = Role.objects.all().order_by('code')
    
    if search_query:
        roles = roles.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # 为每个角色添加用户数量
    roles_with_count = []
    for role in roles:
        role.user_count = User.objects.filter(role=role.code).count()
        roles_with_count.append(role)
    
    # 分页
    from django.core.paginator import Paginator
    paginator = Paginator(roles_with_count, 10)  # 每频10个角色
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'users/role_list.html', {
        'roles': page_obj,
        'search_query': search_query,
        'is_paginated': page_obj.has_other_pages(),
        'page_obj': page_obj,
    })


@login_required
@admin_required
def role_create(request):
    """创建角色（仅系统管理员）"""
    if request.user.role != 'admin':
        messages.error(request, '权限不足：只有系统管理员才能创建角色！')
        return redirect('index')
    
    if request.method == 'POST':
        form = RoleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '角色创建成功！')
            return redirect('users:role_list')
    else:
        form = RoleForm()
    
    return render(request, 'users/role_form.html', {
        'form': form,
        'title': '添加角色'
    })


@login_required
@admin_required
def role_update(request, pk):
    """更新角色（仅系统管理员）"""
    if request.user.role != 'admin':
        messages.error(request, '权限不足：只有系统管理员才能编辑角色！')
        return redirect('index')
    
    role = get_object_or_404(Role, pk=pk)
    
    if request.method == 'POST':
        form = RoleForm(request.POST, instance=role)
        if form.is_valid():
            form.save()
            messages.success(request, '角色信息更新成功！')
            return redirect('users:role_list')
    else:
        form = RoleForm(instance=role)
    
    return render(request, 'users/role_form.html', {
        'form': form,
        'title': '编辑角色',
        'role': role
    })


@login_required
@admin_required
def role_detail(request, pk):
    """角色详情（仅系统管理员）"""
    if request.user.role != 'admin':
        messages.error(request, '权限不足：只有系统管理员才能查看角色详情！')
        return redirect('index')
    
    role = get_object_or_404(Role, pk=pk)
    # 获取该角色的用户
    users = User.objects.filter(role=role.code).order_by('username')
    
    # 获取权限名称
    permission_dict = dict(Role.PERMISSION_CHOICES)
    permissions_display = [permission_dict.get(p, p) for p in role.permissions]
    
    return render(request, 'users/role_detail.html', {
        'role': role,
        'users': users,
        'permissions_display': permissions_display,
    })


@login_required
@admin_required
def role_delete(request, pk):
    """删除角色（仅系统管理员）"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'message': '权限不足！'})
    
    role = get_object_or_404(Role, pk=pk)
    
    # 检查是否有用户使用此角色
    user_count = User.objects.filter(role=role.code).count()
    if user_count > 0:
        return JsonResponse({
            'success': False, 
            'message': f'该角色下还有 {user_count} 个用户，无法删除！'
        })
    
    role_name = role.name
    role.delete()
    
    return JsonResponse({
        'success': True, 
        'message': f'角色 {role_name} 删除成功！'
    })

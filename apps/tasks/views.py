"""
试验任务管理视图
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.http import Http404, JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from datetime import datetime
import json
from django.contrib.auth import get_user_model
User = get_user_model()
from .models import TestTask, TestType, TestTypeField, PriorityType, TaskStatus, SubTask, SubTaskData, TestTaskReport, TestTaskProcessHistory
from .forms import TestTaskForm, TestTaskSearchForm, TestTypeForm, TestTypeFieldForm, TaskDecomposeForm, SubTaskForm, SubTaskDataForm
from .generic_forms import create_generic_test_data_formset


@login_required
def task_list(request):
    """试验任务列表"""
    search_form = TestTaskSearchForm(request.GET)
    tasks = TestTask.objects.select_related(
        'test_type', 'priority', 'status', 'requester', 'assignee'
    ).all()
    
    # 针对试验室主任（manager）的默认视图优化
    if request.user.role == 'manager' and not request.GET.get('status'):
        tasks = tasks.filter(status__code__in=['pending', 'in_progress', 'pending_review'])
    
    # 如果是访客角色，只显示自己申请的任务
    if request.user.role == 'guest':
        tasks = tasks.filter(requester=request.user)
    
    # 如果是试验工程师，只显示分配给自己的任务
    if request.user.role == 'engineer':
        tasks = tasks.filter(assignee=request.user)
    
    # 搜索过滤
    if search_form.is_valid():
        search_query = search_form.cleaned_data.get('search')
        test_type = search_form.cleaned_data.get('test_type')
        priority = search_form.cleaned_data.get('priority')
        status = search_form.cleaned_data.get('status')
        assignee = search_form.cleaned_data.get('assignee')
        date_from = search_form.cleaned_data.get('date_from')
        date_to = search_form.cleaned_data.get('date_to')
        
        if search_query:
            tasks = tasks.filter(
                Q(task_number__icontains=search_query) |
                Q(task_name__icontains=search_query) |
                Q(description__icontains=search_query)
            )
        
        if test_type:
            tasks = tasks.filter(test_type=test_type)
        
        if priority:
            tasks = tasks.filter(priority=priority)
        
        if status:
            tasks = tasks.filter(status=status)
        
        if assignee:
            tasks = tasks.filter(assignee=assignee)
        
        if date_from:
            tasks = tasks.filter(start_date__gte=date_from)
        
        if date_to:
            tasks = tasks.filter(end_date__lte=date_to)
    
    # 试验工程师：优先显示紧急且快到期的任务
    if request.user.role == 'engineer':
        from django.db.models import Case, When, Value, IntegerField
        
        today = timezone.now().date()
        tasks = tasks.annotate(
            is_urgent=Case(
                When(priority__code='high', status__code__in=['pending', 'in_progress'], then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            ),
            is_expiring=Case(
                When(end_date__lte=today, status__code__in=['pending', 'in_progress'], then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            ),
        ).order_by('-is_urgent', '-is_expiring', '-created_at')
    else:
        # 默认按创建时间倒序
        tasks = tasks.order_by('-created_at')
        
    paginator = Paginator(tasks, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_form': search_form,
    }
    return render(request, 'tasks/task_list.html', context)


@login_required
def task_create(request):
    """创建试验任务"""
    if request.method == 'POST':
        form = TestTaskForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            task = form.save(commit=False)
            task.requester = request.user
            # 设置初始状态为待分配
            pending_status = TaskStatus.objects.get(code='pending')
            task.status = pending_status
            
            # 生成任务编号
            today_str = datetime.now().strftime('%Y%m%d')
            count = TestTask.objects.filter(created_at__date=datetime.now().date()).count() + 1
            task.task_number = f"TASK-{today_str}-{count:03d}"
            
            task.save()
            messages.success(request, f'任务 {task.task_number} 创建成功！')
            return redirect('tasks:task_list')
    else:
        form = TestTaskForm(user=request.user)
    
    return render(request, 'tasks/task_form.html', {'form': form, 'title': '创建试验任务'})


@login_required
def task_detail(request, task_id):
    """任务详情"""
    task = get_object_or_404(TestTask, id=task_id)
    
    # 权限检查
    # 访客只能看自己的任务
    if request.user.role == 'guest' and task.requester != request.user:
        messages.error(request, '您没有权限查看此任务！')
        return redirect('tasks:task_list')
    
    # 试验工程师只能看分配给自己的任务
    if request.user.role == 'engineer' and task.assignee != request.user:
        # 但如果是自己申请的也可以看
        if task.requester != request.user:
            messages.error(request, '您没有权限查看此任务！')
            return redirect('tasks:task_list')
            
    reports = task.reports.all().order_by('-created_at')
    statuses = TaskStatus.objects.all()
    users = User.objects.filter(role='engineer')
    
    context = {
        'task': task,
        'reports': reports,
        'statuses': statuses,
        'users': users,
    }
    return render(request, 'tasks/task_detail.html', context)


@login_required
def task_edit(request, task_id):
    """编辑任务"""
    task = get_object_or_404(TestTask, id=task_id)
    
    # 权限检查
    if not (request.user == task.requester or request.user.role in ['manager', 'admin']):
        messages.error(request, '您没有权限编辑此任务！')
        return redirect('tasks:task_detail', task_id=task.id)
    
    # 状态检查：只有待处理状态可以编辑
    if task.status.code != 'pending' and request.user.role not in ['manager', 'admin']:
        messages.error(request, '任务已开始处理，无法编辑！')
        return redirect('tasks:task_detail', task_id=task.id)
        
    if request.method == 'POST':
        form = TestTaskForm(request.POST, request.FILES, instance=task, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, '任务更新成功！')
            return redirect('tasks:task_detail', task_id=task.id)
    else:
        form = TestTaskForm(instance=task, user=request.user)
        
    return render(request, 'tasks/task_form.html', {'form': form, 'title': '编辑试验任务'})


@login_required
def task_delete(request, task_id):
    """删除任务"""
    task = get_object_or_404(TestTask, id=task_id)
    
    # 权限检查
    if not (request.user == task.requester or request.user.role in ['manager', 'admin']):
        messages.error(request, '您没有权限删除此任务！')
        return redirect('tasks:task_detail', task_id=task.id)
        
    if request.method == 'POST':
        task.delete()
        messages.success(request, '任务已删除！')
        return redirect('tasks:task_list')
        
    return render(request, 'confirm_delete.html', {'object': task, 'type': '试验任务'})


@login_required
@require_http_methods(["POST"])
def task_assign(request, task_id):
    """分配任务"""
    # 只有管理员和主任可以分配
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, '权限不足！')
        return redirect('tasks:task_detail', task_id=task_id)
        
    task = get_object_or_404(TestTask, id=task_id)
    assignee_id = request.POST.get('assignee')
    
    if assignee_id:
        assignee = get_object_or_404(User, id=assignee_id)
        task.assignee = assignee
        # 如果当前状态是待分配，自动更新为进行中
        if task.status.code == 'pending':
            in_progress = TaskStatus.objects.get(code='in_progress')
            task.status = in_progress
            task.actual_start_date = timezone.now().date()
        
        task.save()
        messages.success(request, f'任务已分配给 {assignee.username}')
    else:
        task.assignee = None
        task.save()
        messages.info(request, '已取消任务分配')
        
    return redirect('tasks:task_detail', task_id=task_id)


@login_required
@require_http_methods(["POST"])
def task_status_update(request, task_id):
    """更新任务状态"""
    task = get_object_or_404(TestTask, id=task_id)
    
    # 权限检查
    can_update = False
    if request.user.role in ['manager', 'admin']:
        can_update = True
    elif request.user == task.assignee:
        can_update = True
    elif request.user == task.requester and task.status.code == 'pending':
        can_update = True # 申请人可以取消任务
        
    if not can_update:
        return JsonResponse({'success': False, 'message': '权限不足！'})
    
    import json
    data = json.loads(request.body)
    status_id = data.get('status_id')
    remarks = data.get('remarks', '')
    rejection_reason = data.get('rejection_reason', '')
    
    if not status_id:
        return JsonResponse({'success': False, 'message': '状态ID不能为空'})
        
    try:
        new_status = TaskStatus.objects.get(id=status_id)
        old_status_code = task.status.code
        task.status = new_status
        
        # 状态变更逻辑处理
        if new_status.code == 'in_progress' and not task.actual_start_date:
            task.actual_start_date = timezone.now().date()
            
        if new_status.code in ['completed', 'reviewed', 'cancelled'] and not task.actual_end_date:
            task.actual_end_date = timezone.now().date()
            
        # 如果是从进行中变更为待审核（工程师提交）
        if old_status_code == 'in_progress' and new_status.code == 'pending_review':
            task.actual_end_date = timezone.now().date() # 记录提交时间
            
        # 如果是被拒绝或退回
        if rejection_reason:
            task.rejection_reason = rejection_reason
            
        # 如果是从待审核退回到进行中（审核不通过）
        if old_status_code == 'pending_review' and new_status.code == 'in_progress':
            task.actual_end_date = None # 清除结束时间，重新计时
            
        task.save()
        
        return JsonResponse({
            'success': True, 
            'message': f'状态已更新为：{new_status.name}',
            'new_status': new_status.name,
            'new_status_code': new_status.code
        })
    except TaskStatus.DoesNotExist:
        return JsonResponse({'success': False, 'message': '状态不存在'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def task_dashboard(request):
    """任务仪表板"""
    # 统计数据
    total_tasks = TestTask.objects.count()
    pending_tasks = TestTask.objects.filter(status__code='pending').count()
    in_progress_tasks = TestTask.objects.filter(status__code='in_progress').count()
    completed_tasks = TestTask.objects.filter(status__code__in=['completed', 'reviewed']).count()
    
    # 最近的任务
    recent_tasks = TestTask.objects.order_by('-created_at')[:5]
    
    # 用户任务统计
    user_stats = {
        'total': TestTask.objects.filter(requester=request.user).count(),
        'pending': TestTask.objects.filter(requester=request.user, status__code='pending').count(),
        'processing': TestTask.objects.filter(requester=request.user, status__code='in_progress').count(),
        'completed': TestTask.objects.filter(requester=request.user, status__code__in=['completed', 'reviewed']).count(),
    }
    
    # 工程师任务统计
    engineer_stats = None
    if request.user.role == 'engineer':
        engineer_stats = {
            'assigned': TestTask.objects.filter(assignee=request.user).count(),
            'pending': TestTask.objects.filter(assignee=request.user, status__code='in_progress').count(),
            'completed': TestTask.objects.filter(assignee=request.user, status__code__in=['completed', 'reviewed']).count(),
        }
        
    context = {
        'total_tasks': total_tasks,
        'pending_tasks': pending_tasks,
        'in_progress_tasks': in_progress_tasks,
        'completed_tasks': completed_tasks,
        'recent_tasks': recent_tasks,
        'user_stats': user_stats,
        'engineer_stats': engineer_stats,
    }
    return render(request, 'tasks/dashboard.html', context)


@login_required
def my_tasks(request):
    """我的任务"""
    # 重定向到列表页，带上筛选参数
    if request.user.role == 'engineer':
        return redirect('/tasks/?assignee=' + str(request.user.id))
    else:
        return redirect('/tasks/?requester=' + str(request.user.id))


# ==================== 试验类型管理 ====================

@login_required
def test_type_list(request):
    """试验类型列表"""
    # 只有管理员和主任可以管理试验类型
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, '权限不足！')
        return redirect('tasks:dashboard')
        
    search_query = request.GET.get('search', '')
    
    # 基础查询：所有试验类型
    queryset = TestType.objects.all()
    
    # 搜索过滤
    if search_query:
        queryset = queryset.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # 排序：总类型在前，分类型在后；同级按代码排序
    queryset = queryset.order_by('level_type', 'code')
    
    # 统计任务数量 (优化性能：使用annotate)
    from django.db.models import Count
    queryset = queryset.annotate(task_count=Count('testtask'))
    
    # 分页
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'test_types': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'search_query': search_query,
    }
    
    return render(request, 'tasks/test_type_list.html', context)


@login_required
def test_type_create(request):
    """创建试验类型"""
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, '权限不足！')
        return redirect('tasks:dashboard')
        
    if request.method == 'POST':
        form = TestTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '试验类型创建成功！')
            return redirect('tasks:test_type_list')
    else:
        form = TestTypeForm()
        
    return render(request, 'tasks/test_type_form.html', {'form': form, 'title': '创建试验类型'})


@login_required
def test_type_update(request, pk):
    """更新试验类型"""
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, '权限不足！')
        return redirect('tasks:dashboard')
        
    test_type = get_object_or_404(TestType, pk=pk)
    
    if request.method == 'POST':
        form = TestTypeForm(request.POST, instance=test_type)
        if form.is_valid():
            form.save()
            messages.success(request, '试验类型更新成功！')
            return redirect('tasks:test_type_list')
    else:
        form = TestTypeForm(instance=test_type)
        
    return render(request, 'tasks/test_type_form.html', {'form': form, 'title': '更新试验类型'})


@login_required
def test_type_delete(request, pk):
    """删除试验类型"""
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, '权限不足！')
        return redirect('tasks:dashboard')
        
    test_type = get_object_or_404(TestType, pk=pk)
    
    if request.method == 'POST':
        test_type.delete()
        messages.success(request, '试验类型已删除！')
        return redirect('tasks:test_type_list')
        
    return render(request, 'confirm_delete.html', {'object': test_type, 'type': '试验类型'})


@login_required
def test_type_detail(request, pk):
    """试验类型详情（字段配置）"""
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, '权限不足！')
        return redirect('tasks:dashboard')
        
    test_type = get_object_or_404(TestType, pk=pk)
    queryset = test_type.custom_fields.all().order_by('order')
    
    # 分页处理
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('field_page')
    custom_fields = paginator.get_page(page_number)
    
    return render(request, 'tasks/test_type_detail.html', {
        'test_type': test_type,
        'custom_fields': custom_fields
    })


@login_required
def test_type_field_create(request, test_type_id):
    """创建字段配置"""
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, '权限不足！')
        return redirect('tasks:dashboard')
        
    test_type = get_object_or_404(TestType, pk=test_type_id)
    
    if request.method == 'POST':
        form = TestTypeFieldForm(request.POST)
        if form.is_valid():
            field = form.save(commit=False)
            field.test_type = test_type
            field.save()
            messages.success(request, '字段配置已添加！')
            return redirect('tasks:test_type_detail', pk=test_type.id)
    else:
        form = TestTypeFieldForm()
        
    return render(request, 'tasks/test_type_field_form.html', {
        'form': form, 
        'test_type': test_type,
        'title': '添加字段配置'
    })


@login_required
def test_type_field_edit(request, field_id):
    """编辑字段配置"""
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, '权限不足！')
        return redirect('tasks:dashboard')
        
    field = get_object_or_404(TestTypeField, id=field_id)
    
    if request.method == 'POST':
        form = TestTypeFieldForm(request.POST, instance=field)
        if form.is_valid():
            form.save()
            messages.success(request, '字段配置已更新！')
            return redirect('tasks:test_type_detail', pk=field.test_type.id)
    else:
        form = TestTypeFieldForm(instance=field)
        
    return render(request, 'tasks/test_type_field_form.html', {
        'form': form, 
        'test_type': field.test_type,
        'title': '编辑字段配置'
    })


@login_required
def test_type_field_delete(request, field_id):
    """删除字段配置"""
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, '权限不足！')
        return redirect('tasks:dashboard')
        
    field = get_object_or_404(TestTypeField, id=field_id)
    test_type_id = field.test_type.id
    
    if request.method == 'POST':
        field.delete()
        messages.success(request, '字段配置已删除！')
        
    return redirect('tasks:test_type_detail', pk=test_type_id)


@login_required
def test_type_field_image_delete(request, field_id):
    """删除字段配置中的示意图"""
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, '权限不足！')
        return redirect('tasks:dashboard')
        
    field = get_object_or_404(TestTypeField, id=field_id)
    
    if request.method == 'POST':
        # 这里需要实现删除图片的逻辑，目前模型中没有image字段，可能是后续扩展
        # 假设field_options中有image_url
        options = field.field_options
        if 'image_url' in options:
            del options['image_url']
            field.field_options = options
            field.save()
            messages.success(request, '示意图已删除！')
        
    return redirect('tasks:test_type_field_edit', field_id=field.id)


def get_sub_test_types(request):
    """API: 获取子试验类型"""
    parent_id = request.GET.get('parent_id')
    if parent_id:
        sub_types = TestType.objects.filter(parent_id=parent_id).values('id', 'name')
        return JsonResponse(list(sub_types), safe=False)
    return JsonResponse([], safe=False)


# ==================== 子任务管理 ====================

@login_required
def task_decompose(request, task_id):
    """任务分解 - 创建子任务"""
    task = get_object_or_404(TestTask, id=task_id)
    
    # 权限检查
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, '只有试验室主任可以分解任务！')
        return redirect('tasks:task_detail', task_id=task.id)
    
    if request.method == 'POST':
        form = TaskDecomposeForm(request.POST)
        if form.is_valid():
            # 创建子任务
            subtask_count = int(form.cleaned_data['subtask_count'])
            # 这里可以添加更复杂的逻辑，根据选择的试验类型创建不同的子任务
            
            # 为了简化，这里先重定向到子任务列表，实际应该在页面上动态添加子任务表单
            messages.info(request, '功能开发中...')
            return redirect('tasks:task_detail', task_id=task.id)
    else:
        form = TaskDecomposeForm()
        
    # 获取已有的子任务
    subtasks = task.subtasks.all()
    
    # 处理添加子任务的请求
    if 'add_subtask' in request.POST:
        sub_form = SubTaskForm(request.POST)
        if sub_form.is_valid():
            subtask = sub_form.save(commit=False)
            subtask.parent_task = task
            subtask.status = TaskStatus.objects.get(code='pending')
            # 生成子任务编号
            count = task.subtasks.count() + 1
            subtask.subtask_number = f"{task.task_number}-SUB{count:02d}"
            subtask.save()
            messages.success(request, '子任务添加成功！')
            return redirect('tasks:task_decompose', task_id=task.id)
    else:
        sub_form = SubTaskForm(initial={'test_type': task.test_type})
    
    return render(request, 'tasks/task_decompose.html', {
        'task': task,
        'subtasks': subtasks,
        'sub_form': sub_form
    })


@login_required
def subtask_detail(request, subtask_id):
    """子任务详情"""
    subtask = get_object_or_404(SubTask, id=subtask_id)
    
    # 尝试获取试验数据对象
    try:
        data_instance = subtask.test_data
    except SubTaskData.DoesNotExist:
        data_instance = None
        
    # 获取试验数据列表 (JSON)
    test_data_list = []
    if data_instance:
        raw_data = data_instance.test_data
        if isinstance(raw_data, list):
            test_data_list = raw_data
        elif raw_data:
            test_data_list = [raw_data]
            
    # 获取字段配置
    all_fields = []
    batch_fields = []
    meta_fields = []
    
    if subtask.test_type:
        if hasattr(subtask.test_type, 'get_all_fields'):
             all_fields = subtask.test_type.get_all_fields()
        else:
             all_fields = list(subtask.test_type.custom_fields.filter(is_active=True).order_by('order'))
        
        # 分类字段
        for field in all_fields:
            if field.is_batch_input_enabled:
                batch_fields.append(field)
            else:
                meta_fields.append(field)
    
    # 实例化元数据表单（用于在详情页直接编辑）
    meta_form = None
    if meta_fields or data_instance:
        meta_form = SubTaskDataForm(instance=data_instance, subtask=subtask)
        
    return render(request, 'tasks/subtask_detail.html', {
        'subtask': subtask,
        'data_instance': data_instance, # SubTaskData 对象
        'test_data_list': test_data_list, # 批量录入的数据列表
        'all_fields': all_fields,
        'batch_fields': batch_fields,
        'meta_fields': meta_fields,
        'meta_form': meta_form
    })


@login_required
def subtask_delete(request, subtask_id):
    """删除子任务"""
    subtask = get_object_or_404(SubTask, id=subtask_id)
    task_id = subtask.parent_task.id
    
    if request.user.role not in ['manager', 'admin']:
        messages.error(request, '权限不足！')
        return redirect('tasks:task_detail', task_id=task_id)
        
    if request.method == 'POST':
        subtask.delete()
        messages.success(request, '子任务已删除！')
        
    return redirect('tasks:task_decompose', task_id=task_id)


@login_required
def subtask_data_edit(request, subtask_id):
    """编辑子任务数据（通用）"""
    subtask = get_object_or_404(SubTask, id=subtask_id)
    
    # 检查是否已有数据
    try:
        data_instance = subtask.test_data
    except SubTaskData.DoesNotExist:
        data_instance = None
        
    if request.method == 'POST':
        form = SubTaskDataForm(request.POST, request.FILES, instance=data_instance)
        if form.is_valid():
            data = form.save(commit=False)
            data.subtask = subtask
            data.save()
            
            # 更新子任务状态
            if subtask.status.code == 'pending':
                subtask.status = TaskStatus.objects.get(code='in_progress')
                subtask.actual_start_date = timezone.now().date()
                subtask.save()
                
            messages.success(request, '试验数据已保存！')
            return redirect('tasks:subtask_detail', subtask_id=subtask.id)
    else:
        form = SubTaskDataForm(instance=data_instance)
        
    return render(request, 'tasks/subtask_data_form.html', {
        'form': form,
        'subtask': subtask
    })


# ==================== 通用批量试验数据管理 ====================

@login_required
def generic_test_data_entry(request, subtask_id):
    """通用试验数据批量录入"""
    # 优化查询：预加载关联的试验类型
    subtask = get_object_or_404(SubTask.objects.select_related('test_type'), id=subtask_id)
    test_type = subtask.test_type
    
    # 权限检查
    if not (request.user == subtask.assignee or request.user.role in ['manager', 'admin']):
        messages.error(request, '权限不足！')
        return redirect('tasks:subtask_detail', subtask_id=subtask.id)
    
    # 预加载字段配置（性能优化）
    all_fields = []
    if hasattr(test_type, 'get_all_fields'):
        all_fields = test_type.get_all_fields()
    else:
        all_fields = list(test_type.custom_fields.filter(is_active=True).order_by('order', 'id'))
        
    # 仅保留启用批量录入的字段
    fields = [f for f in all_fields if f.is_batch_input_enabled]
    
    # 获取或创建 SubTaskData
    subtask_data, created = SubTaskData.objects.get_or_create(subtask=subtask)
    
    # 获取已有数据
    existing_data = subtask_data.test_data
    if not isinstance(existing_data, list):
        existing_data = [existing_data] if existing_data else []
        
    # 处理Excel导入
    imported_data = []
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        # 简单的Excel读取逻辑
        import openpyxl
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # 读取表头 - 增强处理：去除星号和首尾空格
            headers = []
            for cell in ws[1]:
                val = str(cell.value).strip() if cell.value is not None else ""
                # 去掉末尾的星号（如果是必填项标记）
                if val.endswith('*'):
                    val = val[:-1]
                headers.append(val)
            
            # 读取数据
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row): continue  # 跳过空行
                row_data = {}
                # 尝试匹配字段
                for field in fields:
                    if field.field_name in headers:
                        idx = headers.index(field.field_name)
                        if idx < len(row):
                            val = row[idx]
                            # 处理日期时间类型
                            if val is not None:
                                # 注意：这里不要导入 datetime 类，因为函数外层已经导入了 datetime 模块
                                # 最好使用 type() 或 isinstance 检查
                                from datetime import date as dt_date, datetime as dt_datetime
                                if isinstance(val, (dt_date, dt_datetime)):
                                    # 如果是日期对象，根据字段类型转换
                                    if field.field_type == 'date':
                                        val = val.strftime('%Y-%m-%d')
                                    elif field.field_type == 'datetime':
                                        val = val.strftime('%Y-%m-%dT%H:%M')
                                    else:
                                        val = str(val)
                            
                            row_data[field.field_code] = val
                imported_data.append(row_data)
                
            if imported_data:
                messages.success(request, f'成功从 Excel 导入 {len(imported_data)} 条数据。')
            else:
                messages.warning(request, 'Excel 文件中没有发现有效数据。')
        except Exception as e:
            messages.error(request, f'Excel 文件解析失败：{str(e)}')
    
    # 准备初始数据
    initial_data = existing_data + imported_data
    
    extra_rows = 1 if not initial_data else 0
    FormSet = create_generic_test_data_formset(test_type, extra=extra_rows)
    
    # 准备表单参数
    form_kwargs = {
        'test_type': test_type,
        'subtask': subtask,
        'fields_config': fields
    }
    
    # 如果是POST且不是文件上传，或者是文件上传后的重新渲染
    if request.method == 'POST' and not request.FILES.get('excel_file'):
        formset = FormSet(request.POST, request.FILES, form_kwargs=form_kwargs)
        
        if formset.is_valid():
            new_data_list = []
            count = 0
            
            # 处理表单数据
            for form in formset:
                # 跳过空表单 (未修改且无初始值)
                if not form.has_changed() and not form.initial:
                    continue
                
                # 获取数据
                try:
                    row_data = form.get_test_data()
                except AttributeError:
                    # 如果form没有get_test_data方法（可能未正确初始化），手动获取
                    row_data = form.cleaned_data
                
                if not row_data:
                    continue
                    
                # 确保有试验编号
                if not row_data.get('test_number'):
                    today_str = datetime.now().strftime('%Y%m%d')
                    prefix = test_type.code[:2].upper()
                    row_data['test_number'] = f"{prefix}-{today_str}-{count + 1:04d}"
                
                # 添加或保留创建时间
                if not row_data.get('created_at'):
                    row_data['created_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                new_data_list.append(row_data)
                count += 1
            
            # 保存到 SubTaskData
            subtask_data.test_data = new_data_list
            subtask_data.save()
            
            # 更新子任务状态
            if subtask.status.code == 'pending':
                subtask.status = TaskStatus.objects.get(code='in_progress')
                subtask.actual_start_date = timezone.now().date()
                subtask.save()
                
            messages.success(request, f'成功保存 {count} 条试验数据！')
            return redirect('tasks:generic_test_data_list', subtask_id=subtask.id)
        else:
            messages.error(request, '数据验证失败，请检查输入。')
    else:
        # 初始化表单集
        formset = FormSet(initial=initial_data, form_kwargs=form_kwargs)
    
    # 准备字段配置的JSON数据
    fields_json = []
    for field in fields:
        field_data = {
            'code': field.field_code,
            'name': field.field_name,
            'type': field.field_type,
            'required': field.is_required,
            'options': field.field_options
        }
        fields_json.append(field_data)
        
    context = {
        'subtask': subtask,
        'formset': formset,
        'fields': fields,
        'fields_json': json.dumps(fields_json),
    }
    return render(request, 'tasks/generic_test_data_entry.html', context)


@login_required
def generic_test_data_list(request, subtask_id):
    """通用试验数据列表"""
    subtask = get_object_or_404(SubTask, id=subtask_id)
    
    # 获取数据
    if hasattr(subtask, 'test_data'):
        data_list = subtask.test_data.test_data
        if not isinstance(data_list, list):
            data_list = [data_list] if data_list else []
    else:
        data_list = []
    
    # 获取字段配置
    test_type = subtask.test_type
    if hasattr(test_type, 'get_all_fields'):
        all_fields = test_type.get_all_fields()
    else:
        all_fields = list(test_type.custom_fields.filter(is_active=True).order_by('order'))
        
    fields = [f for f in all_fields if f.is_batch_input_enabled]

    # 导出功能
    if request.GET.get('export') == 'true':
        from openpyxl import Workbook
        from django.http import HttpResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "试验数据"
        
        # 写入表头
        header = ['试验编号'] + [f.field_name for f in fields] + ['录入时间']
        ws.append(header)
        
        # 写入数据
        for data in data_list:
            row = [data.get('test_number', '')]
            for f in fields:
                val = data.get(f.field_code, '')
                # 处理特殊类型
                if f.field_type == 'checkbox':
                    val = '是' if val in [True, 'True', 'true', '1'] else '否'
                elif f.field_type == 'file':
                    if isinstance(val, list):
                        val = f"包含 {len(val)} 个文件"
                    elif val:
                        val = "包含 1 个文件"
                    else:
                        val = ""
                row.append(str(val) if val is not None else '')
            row.append(data.get('created_at', ''))
            ws.append(row)
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"{subtask.subtask_number}_data_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    # 分页处理
    paginator = Paginator(data_list, 10) # 每页10条
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'tasks/generic_test_data_list.html', {
        'subtask': subtask,
        'page_obj': page_obj, # 替换 data_list
        'data_list': page_obj, # 保持兼容
        'fields': fields
    })


@login_required
def generic_test_data_delete(request, data_id):
    """删除通用试验数据"""
    data = get_object_or_404(GenericTestData, id=data_id)
    subtask = data.subtask
    
    if not (request.user == subtask.assignee or request.user.role in ['manager', 'admin']):
        messages.error(request, '权限不足！')
        return redirect('tasks:generic_test_data_list', subtask_id=subtask.id)
        
    if request.method == 'POST':
        data.delete()
        messages.success(request, '数据已删除！')
        
    return redirect('tasks:generic_test_data_list', subtask_id=subtask.id)


@login_required
def download_generic_test_template(request, subtask_id):
    """下载通用试验数据导入模板"""
    subtask = get_object_or_404(SubTask, id=subtask_id)
    test_type = subtask.test_type
    
    if hasattr(test_type, 'get_all_fields'):
        all_fields = test_type.get_all_fields()
    else:
        all_fields = list(test_type.custom_fields.filter(is_active=True).order_by('order'))
        
    fields = [f for f in all_fields if f.is_batch_input_enabled]
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据导入模板"
    
    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 写入表头
    for col_num, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = field.field_name
        cell.font = header_font
        cell.fill = header_fill
        
        # 添加批注/提示（可选）
        
    # 自动调整列宽
    for column in ws.columns:
        length = len(str(column[0].value))
        ws.column_dimensions[column[0].column_letter].width = length + 5
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{test_type.name}_导入模板.xlsx"
    # 处理中文文件名
    from django.utils.encoding import escape_uri_path
    response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(filename)}"'
    
    wb.save(response)
    return response


class TestDataWrapper:
    """试验数据包装器，用于统一不同类型的试验数据展示"""
    def __init__(self, data_obj, data_type):
        self.data_obj = data_obj
        self.data_type = data_type  # 'generic' 或 'disc_spring'
    
    @property
    def id(self):
        return self.data_obj.id
    
    @property
    def test_number(self):
        if self.data_type == 'generic':
            return self.data_obj.test_number
        elif self.data_type == 'disc_spring':
            return self.data_obj.test_number
    
    @property
    def subtask(self):
        return self.data_obj.subtask
    
    @property
    def created_at(self):
        return self.data_obj.created_at
    
    @property
    def updated_at(self):
        return self.data_obj.updated_at
    
    @property
    def test_data(self):
        if self.data_type == 'generic':
            return self.data_obj.test_data
        elif self.data_type == 'disc_spring':
            # 将碟簧数据转换为字典格式
            return {
                'batch_number': self.data_obj.batch_number,
                'batch_total': self.data_obj.batch_total,
                'test_quantity': self.data_obj.test_quantity,
                'supplier': self.data_obj.supplier,
                'disc_spring_model': self.data_obj.disc_spring_model,
                'application_date': self.data_obj.application_date,
                'displacement_2mm': self.data_obj.displacement_2mm,
                'displacement_2_5mm': self.data_obj.displacement_2_5mm,
                'inspector_name': self.data_obj.inspector_name,
                'inspection_time': self.data_obj.inspection_time,
                'remarks': self.data_obj.remarks,
            }


@login_required
def test_data_list(request):
    """试验数据管理 - 任务查询"""
    # 获取筛选参数
    main_type_id = request.GET.get('main_type_id')
    sub_type_id = request.GET.get('sub_type_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    keyword = request.GET.get('keyword')
    
    # 确定查询模式：如果选择了子任务类型，则查询子任务；否则查询主任务
    mode = 'subtask' if sub_type_id else 'maintask'
    
    # 获取试验类型下拉选项
    main_types = TestType.objects.filter(level_type=1)
    sub_types = TestType.objects.filter(level_type=2)
    # 获取任务状态选项
    statuses = TaskStatus.objects.all().order_by('id')
    
    data_list = []
    
    if mode == 'maintask':
        # 查询主任务
        queryset = TestTask.objects.select_related(
            'test_type', 'status', 'priority', 'requester', 'assignee'
        ).all()
        
        # 筛选条件
        if main_type_id:
            queryset = queryset.filter(test_type_id=main_type_id)
            
        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)
            
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)
            
        if keyword:
            queryset = queryset.filter(
                Q(task_number__icontains=keyword) | 
                Q(task_name__icontains=keyword)
            )
        
        # 状态筛选
        status_id = request.GET.get('status_id')
        if status_id:
            queryset = queryset.filter(status_id=status_id)
            
        # 默认查询条件：所有未完成的任务（不限制月份，避免跨月无数据）
        # 注意：如果用户主动选择了状态，就不应用默认排除已完成的逻辑
        if not any([main_type_id, start_date, end_date, keyword, status_id]):
            queryset = queryset.exclude(status__code__in=['completed', 'reviewed', 'cancelled'])
            
        # 预加载报告和文件信息，优化查询
        queryset = queryset.prefetch_related('reports')
        
        data_list = queryset.order_by('-created_at')
        
    else:
        # 查询子任务
        queryset = SubTask.objects.select_related(
            'test_type', 'status', 'parent_task', 'assignee'
        ).all()
        
        if sub_type_id:
            queryset = queryset.filter(test_type_id=sub_type_id)
            
        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)
            
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)
            
        if keyword:
            queryset = queryset.filter(
                Q(subtask_number__icontains=keyword) | 
                Q(subtask_name__icontains=keyword) |
                Q(parent_task__task_number__icontains=keyword)
            )

        # 状态筛选
        status_id = request.GET.get('status_id')
        if status_id:
            queryset = queryset.filter(status_id=status_id)
            
        data_list = queryset.order_by('-created_at')

    # 导出功能
    if request.GET.get('export') == 'excel':
        return export_tasks_excel(data_list, mode)

    # 分页处理
    paginator = Paginator(data_list, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    
    context = {
        'page_obj': page_obj,
        'main_types': main_types,
        'sub_types': sub_types,
        'statuses': statuses,
        'mode': mode,
        'filters': {
            'main_type_id': main_type_id,
            'sub_type_id': sub_type_id,
            'start_date': start_date,
            'end_date': end_date,
            'keyword': keyword,
            'status_id': request.GET.get('status_id')
        }
    }
    
    return render(request, 'tasks/test_data_list.html', context)


def export_tasks_excel(queryset, mode):
    """导出任务数据到Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "任务数据"
    
    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    if mode == 'maintask':
        headers = [
            '任务编号', '任务名称', '产品名称', '产品型号', '试验类型', 
            '状态', '优先级', '申请人', '申请人电话', '申请人部门',
            '负责人', '任务描述', '试验过程描述', '试验大纲内容', '试验报告内容',
            '开始日期', '结束日期', '实际开始日期', '实际结束日期', '创建时间'
        ]
        ws.append(headers)
        for task in queryset:
            # 处理富文本内容，去除HTML标签
            process_desc = task.test_process_description
            if process_desc:
                from django.utils.html import strip_tags
                process_desc = strip_tags(process_desc)
                
            row = [
                task.task_number, 
                task.task_name, 
                task.product_name, 
                task.product_model, 
                task.test_type.name, 
                task.status.name,
                task.priority.name, 
                task.requester_name or (task.requester.get_full_name() if task.requester else '') or (task.requester.username if task.requester else ''), 
                task.requester_phone,
                task.requester_department,
                task.assignee.get_full_name() if task.assignee else (task.assignee.username if task.assignee else ''),
                task.description,
                process_desc,
                task.test_outline,
                task.test_report,
                task.start_date.strftime('%Y-%m-%d') if task.start_date else '',
                task.end_date.strftime('%Y-%m-%d') if task.end_date else '',
                task.actual_start_date.strftime('%Y-%m-%d') if task.actual_start_date else '',
                task.actual_end_date.strftime('%Y-%m-%d') if task.actual_end_date else '',
                task.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            ws.append(row)
    else:
        headers = [
            '子任务编号', '子任务名称', '主任务编号', '试验类型', 
            '状态', '负责人', '子任务描述', 
            '开始日期', '结束日期', '实际开始日期', '实际结束日期', '创建时间'
        ]
        ws.append(headers)
        for task in queryset:
            row = [
                task.subtask_number, 
                task.subtask_name, 
                task.parent_task.task_number,
                task.test_type.name, 
                task.status.name,
                task.assignee.get_full_name() if task.assignee else (task.assignee.username if task.assignee else ''),
                task.description,
                task.start_date.strftime('%Y-%m-%d') if task.start_date else '',
                task.end_date.strftime('%Y-%m-%d') if task.end_date else '',
                task.actual_start_date.strftime('%Y-%m-%d') if task.actual_start_date else '',
                task.actual_end_date.strftime('%Y-%m-%d') if task.actual_end_date else '',
                task.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            ws.append(row)
            
    # 设置样式
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"任务导出_{mode}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def test_data_search(request):
    """试验数据搜索 - 专门处理搜索请求"""
    # 重定向到列表页面，带搜索参数
    return redirect('tasks:test_data_list')


@login_required
def test_data_edit(request, data_id):
    """编辑试验数据"""
    # 尝试获取通用试验数据
    try:
        data = get_object_or_404(GenericTestData, id=data_id)
        data_wrapper = TestDataWrapper(data, 'generic')
    except:
        # 如果不是通用试验数据，尝试获取碟簧试验数据
        try:
            data = get_object_or_404(DiscSpringTestData, id=data_id)
            data_wrapper = TestDataWrapper(data, 'disc_spring')
        except:
            # 如果都找不到，返回404
            raise Http404("数据不存在")
    
    # 权限检查
    if not (request.user == data_wrapper.subtask.assignee or request.user.role in ['manager', 'admin']):
        messages.error(request, '权限不足：您没有权限编辑此试验数据！')
        return redirect('tasks:test_data_list')
    
    # 获取试验类型字段配置
    test_type = data_wrapper.subtask.test_type
    # fields = test_type.custom_fields.filter(is_active=True).order_by('order')
    # 使用支持继承的方法
    fields = test_type.get_all_fields()
    
    if request.method == 'POST':
        try:
            if data_wrapper.data_type == 'generic':
                # 更新通用试验数据
                test_data = {}
                # 处理文件上传
                if request.FILES:
                    from django.core.files.storage import default_storage
                    import os
                    
                    for field in fields:
                        if field.field_type == 'file':
                            files = request.FILES.getlist(field.field_code)
                            if files:
                                file_urls = []
                                # 如果已有文件，先保留
                                current_value = test_data.get(field.field_code)
                                if current_value:
                                    # 如果是单个文件路径，转为列表
                                    if isinstance(current_value, str):
                                        file_urls.append(current_value)
                                    # 如果已经是列表（JSON格式），则直接使用
                                    elif isinstance(current_value, list):
                                        file_urls.extend(current_value)
                                
                                for uploaded_file in files:
                                    # 保存文件
                                    file_name = f"test_data/{timezone.now().strftime('%Y%m%d')}/{uploaded_file.name}"
                                    file_path = default_storage.save(file_name, uploaded_file)
                                    file_url = default_storage.url(file_path)
                                    file_urls.append(file_url)
                                
                                # 更新数据：如果是多文件，保存为列表（需前端和模型支持JSON存储或分隔符），如果是单文件字段兼容处理
                                # 考虑到通用数据存储是JSONField，可以直接存列表
                                # 但为了兼容旧的单文件逻辑，如果只有一个文件且之前没有值，可能存为字符串
                                # 这里改为统一存列表，或者用分号分隔的字符串
                                # 为了最小化改动，这里使用分号分隔的字符串来存储多文件路径
                                test_data[field.field_code] = ";".join(file_urls)

                for field in fields:
                    # 跳过文件字段，因为上面已经处理了
                    if field.field_type == 'file':
                        continue
                        
                    value = request.POST.get(field.field_code)
                    if value is not None:
                        if field.field_type == 'number':
                            test_data[field.field_code] = int(value) if value else None
                        elif field.field_type == 'decimal':
                            test_data[field.field_code] = float(value) if value else None
                        elif field.field_type == 'checkbox':
                            test_data[field.field_code] = bool(request.POST.get(field.field_code))
                        else:
                            test_data[field.field_code] = value
                    else:
                        test_data[field.field_code] = None
                
                data.test_data = test_data
                data.updated_at = timezone.now()
                data.save()
                messages.success(request, f'数据 {data.test_number} 更新成功')
            elif data_wrapper.data_type == 'disc_spring':
                # 更新碟簧试验数据
                for field in fields:
                    value = request.POST.get(field.field_code)
                    if hasattr(data, field.field_code) and value is not None:
                        if field.field_type == 'number':
                            setattr(data, field.field_code, int(value) if value else None)
                        elif field.field_type == 'decimal':
                            setattr(data, field.field_code, float(value) if value else None)
                        elif field.field_type == 'date':
                            try:
                                setattr(data, field.field_code, datetime.strptime(value, '%Y-%m-%d').date())
                            except ValueError:
                                pass
                        elif field.field_type == 'datetime':
                            try:
                                setattr(data, field.field_code, datetime.strptime(value, '%Y-%m-%dT%H:%M'))
                            except ValueError:
                                pass
                        elif field.field_type == 'checkbox':
                            setattr(data, field.field_code, bool(request.POST.get(field.field_code)))
                        else:
                            setattr(data, field.field_code, value)
                
                data.updated_at = timezone.now()
                data.save()
                messages.success(request, f'数据 {data.test_number} 更新成功')
            
            return redirect('tasks:test_data_list')
        except Exception as e:
            messages.error(request, f'更新失败：{str(e)}')
    
    context = {
        'data': data_wrapper,
        'test_type': test_type,
        'fields': fields,
    }
    
    return render(request, 'tasks/test_data_edit.html', context)


@login_required
@require_http_methods(["POST"])
def test_data_delete(request, data_id):
    """删除试验数据"""
    # 尝试获取通用试验数据
    try:
        data = get_object_or_404(GenericTestData, id=data_id)
        data_wrapper = TestDataWrapper(data, 'generic')
    except:
        # 如果不是通用试验数据，尝试获取碟簧试验数据
        try:
            data = get_object_or_404(DiscSpringTestData, id=data_id)
            data_wrapper = TestDataWrapper(data, 'disc_spring')
        except:
            # 如果都找不到，返回错误
            return JsonResponse({'success': False, 'message': '数据不存在！'})
    
    # 权限检查
    if not (request.user == data_wrapper.subtask.assignee or request.user.role in ['manager', 'admin']):
        return JsonResponse({'success': False, 'message': '权限不足！'})
    
    try:
        test_number = data_wrapper.test_number
        data.delete()
        return JsonResponse({
            'success': True, 
            'message': f'数据 {test_number} 删除成功'
        })
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'删除失败：{str(e)}'
        })


def export_test_data_excel(data_list):
    """导出试验数据为Excel格式"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from datetime import datetime
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "试验数据"
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 获取所有可能的字段（合并所有试验类型的字段）
    all_fields = {}
    for data in data_list:
        if hasattr(data, 'fields'):
            for field in data.fields.all():
                if field.field_code not in all_fields:
                    all_fields[field.field_code] = field.field_name
    
    # 设置表头
    headers = [
        '序号', '试验编号', '试验类型', '子任务编号', '主任务编号', 
        '试验员', '创建时间'
    ]
    
    # 添加动态字段到表头
    field_headers = list(all_fields.values())
    headers.extend(field_headers)
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入数据
    for row_num, data in enumerate(data_list, 2):
        # 基本信息
        ws.cell(row=row_num, column=1, value=row_num-1)  # 序号
        ws.cell(row=row_num, column=2, value=data.test_number)  # 试验编号
        ws.cell(row=row_num, column=3, value=data.subtask.test_type.name)  # 试验类型
        ws.cell(row=row_num, column=4, value=data.subtask.subtask_number)  # 子任务编号
        ws.cell(row=row_num, column=5, value=data.subtask.parent_task.task_number)  # 主任务编号
        
        # 试验员
        assignee_name = ''
        if data.subtask.assignee:
            assignee_name = data.subtask.assignee.get_full_name() or data.subtask.assignee.username
        ws.cell(row=row_num, column=6, value=assignee_name)
        
        # 创建时间
        ws.cell(row=row_num, column=7, value=data.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        # 动态字段数据
        col_offset = 8  # 从第8列开始写入动态字段数据
        if data.data_type == 'disc_spring':
            # 碟簧试验数据
            field_values = {
                'batch_number': data.test_data.get('batch_number', ''),
                'batch_total': data.test_data.get('batch_total', ''),
                'test_quantity': data.test_data.get('test_quantity', ''),
                'supplier': data.test_data.get('supplier', ''),
                'disc_spring_model': data.test_data.get('disc_spring_model', ''),
                'application_date': data.test_data.get('application_date', ''),
                'displacement_2mm': data.test_data.get('displacement_2mm', ''),
                'displacement_2_5mm': data.test_data.get('displacement_2_5mm', ''),
                'inspector_name': data.test_data.get('inspector_name', ''),
                'inspection_time': data.test_data.get('inspection_time', ''),
                'remarks': data.test_data.get('remarks', ''),
            }
            
            for i, field_code in enumerate(all_fields.keys()):
                value = field_values.get(field_code, '')
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=row_num, column=col_offset+i, value=str(value) if value else '')
        elif data.data_type == 'generic':
            # 通用试验数据
            for i, field_code in enumerate(all_fields.keys()):
                value = data.test_data.get(field_code, '')
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=row_num, column=col_offset+i, value=str(value) if value else '')
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
    
    # 创建响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"试验数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def export_test_data_csv(data_list):
    """导出试验数据为CSV格式"""
    import csv
    from django.http import HttpResponse
    from datetime import datetime
    
    # 创建响应
    response = HttpResponse(content_type='text/csv')
    filename = f"试验数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')  # BOM for Excel
    
    writer = csv.writer(response)
    
    # 获取所有可能的字段（合并所有试验类型的字段）
    all_fields = {}
    for data in data_list:
        if hasattr(data, 'fields'):
            for field in data.fields.all():
                if field.field_code not in all_fields:
                    all_fields[field.field_code] = field.field_name
    
    # 写入表头
    headers = [
        '序号', '试验编号', '试验类型', '子任务编号', '主任务编号', 
        '试验员', '创建时间'
    ]
    
    # 添加动态字段到表头
    field_headers = list(all_fields.values())
    headers.extend(field_headers)
    writer.writerow(headers)
    
    # 写入数据
    for row_num, data in enumerate(data_list, 1):
        # 基本信息
        row_data = [
            row_num,  # 序号
            data.test_number,  # 试验编号
            data.subtask.test_type.name,  # 试验类型
            data.subtask.subtask_number,  # 子任务编号
            data.subtask.parent_task.task_number,  # 主任务编号
        ]
        
        # 试验员
        assignee_name = ''
        if data.subtask.assignee:
            assignee_name = data.subtask.assignee.get_full_name() or data.subtask.assignee.username
        row_data.append(assignee_name)
        
        # 创建时间
        row_data.append(data.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        # 动态字段数据
        if data.data_type == 'disc_spring':
            # 碟簧试验数据
            field_values = {
                'batch_number': data.test_data.get('batch_number', ''),
                'batch_total': data.test_data.get('batch_total', ''),
                'test_quantity': data.test_data.get('test_quantity', ''),
                'supplier': data.test_data.get('supplier', ''),
                'disc_spring_model': data.test_data.get('disc_spring_model', ''),
                'application_date': data.test_data.get('application_date', ''),
                'displacement_2mm': data.test_data.get('displacement_2mm', ''),
                'displacement_2_5mm': data.test_data.get('displacement_2_5mm', ''),
                'inspector_name': data.test_data.get('inspector_name', ''),
                'inspection_time': data.test_data.get('inspection_time', ''),
                'remarks': data.test_data.get('remarks', ''),
            }
            
            for field_code in all_fields.keys():
                value = field_values.get(field_code, '')
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                row_data.append(str(value) if value else '')
        elif data.data_type == 'generic':
            # 通用试验数据
            for field_code in all_fields.keys():
                value = data.test_data.get(field_code, '')
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                row_data.append(str(value) if value else '')
        
        writer.writerow(row_data)
    
    return response


# ==================== 试验任务报告管理 ====================

@login_required
@require_http_methods(["POST"])
def task_report_upload(request, task_id):
    """上传试验报告"""
    task = get_object_or_404(TestTask, id=task_id)
    
    # 权限检查：只有负责人、试验室主管和系统管理员可以上传报告
    if not (request.user == task.assignee or request.user.role in ['manager', 'admin']):
        return JsonResponse({'success': False, 'message': '权限不足：您没有权限上传报告！'})
    
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': '未找到上传文件！'})
    
    file = request.FILES['file']
    
    # 文件大小检查 (50MB)
    if file.size > 50 * 1024 * 1024:
        return JsonResponse({'success': False, 'message': '文件过大，最大支持 50MB！'})
    
    # 文件类型检查
    allowed_types = ['.pdf', '.doc', '.docx', '.ppt', '.pptx', '.xls', '.xlsx', '.txt', '.jpg', '.jpeg', '.png']
    import os
    ext = os.path.splitext(file.name)[1].lower()
    if ext not in allowed_types:
        return JsonResponse({'success': False, 'message': '不支持的文件类型！'})
    
    try:
        report = TestTaskReport.objects.create(
            task=task,
            file=file,
            name=file.name,
            size=file.size,
            uploader=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': '文件上传成功！',
            'report': {
                'id': report.id,
                'name': report.name,
                'size': report.size,
                'url': report.file.url,
                'uploader': report.uploader.username,
                'created_at': report.created_at.strftime('%Y-%m-%d %H:%M')
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'上传失败：{str(e)}'})


@login_required
@require_http_methods(["POST"])
def task_report_delete(request, report_id):
    """删除试验报告"""
    report = get_object_or_404(TestTaskReport, id=report_id)
    
    # 权限检查：只有上传者、试验室主管和系统管理员可以删除
    if not (request.user == report.uploader or request.user.role in ['manager', 'admin']):
        return JsonResponse({'success': False, 'message': '权限不足！'})
    
    try:
        report_name = report.name
        report.delete()
        return JsonResponse({
            'success': True, 
            'message': f'文件 {report_name} 删除成功！'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'删除失败：{str(e)}'})


# ==================== 试验过程描述管理 ====================

@login_required
@require_http_methods(["POST"])
def task_process_update(request, task_id):
    """更新试验过程描述（自动保存）"""
    task = get_object_or_404(TestTask, id=task_id)
    
    # 权限检查：只有负责人、管理员、主管可以编辑
    # 并且任务状态不能是已归档
    if task.status.code == 'archived':
        return JsonResponse({'success': False, 'message': '任务已归档，无法编辑'})

    if not (request.user == task.assignee or request.user.role in ['manager', 'admin']):
        return JsonResponse({'success': False, 'message': '权限不足'})
        
    try:
        data = json.loads(request.body)
        content = data.get('content', '')
        
        # 简单字数校验
        if len(content) > 5000: # 稍微放宽一点后端限制
             return JsonResponse({'success': False, 'message': '内容超过长度限制'})

        # 只有内容变更时才保存
        if content != task.test_process_description:
            # 保存历史记录
            # 只有当上一次历史记录不是同一个人的最近一次（避免频繁自动保存产生垃圾数据）
            # 或者简单点：每次变更都存，但前端debounce控制频率
            TestTaskProcessHistory.objects.create(
                task=task,
                content=content,
                updated_by=request.user
            )
            
            task.test_process_description = content
            task.save(update_fields=['test_process_description', 'updated_at'])
            
        return JsonResponse({'success': True, 'message': '已保存'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def task_process_history(request, task_id):
    """获取试验过程历史记录"""
    task = get_object_or_404(TestTask, id=task_id)
    history_list = task.process_history.select_related('updated_by').all()[:20] # 最近20条
    
    data = []
    for h in history_list:
        data.append({
            'id': h.id,
            'updated_by': h.updated_by.username if h.updated_by else '未知',
            'created_at': h.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'content': h.content  # 返回完整内容以便恢复或查看
        })
        
    return JsonResponse({'success': True, 'history': data})
